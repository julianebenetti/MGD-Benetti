"""
Importador Kalodata -> Supabase (Espião TikTok Shop)

Lê os CSV/XLSX que a Juliane exporta do Kalodata e grava no Supabase, nas tabelas
tiktok_lojas e tiktok_ranking_produtos. Cada import é um snapshot do dia, então dá
pra acompanhar a evolução de GMV semana a semana.

Uso:
    python tiktok-sync.py                          # lê tudo que estiver em dados-kalodata/
    python tiktok-sync.py export.csv               # lê um arquivo específico
    python tiktok-sync.py export.csv --dry-run     # só mostra o mapeamento, não grava
    python tiktok-sync.py export.csv --periodo 30d --data 2026-09-06
    python tiktok-sync.py videos.csv --produto "Calça Pantalona"   # export de Vídeo e Ads

Variáveis de ambiente:
    SUPABASE_URL  (default: projeto da AfiliDash)
    SUPABASE_KEY  (obrigatória — anon key ou service role)

Os números do Kalodata são ESTIMATIVA a partir de sinais públicos do TikTok.
Não são dado oficial da plataforma. A coluna fonte guarda isso.
"""
import csv
import io
import json
import os
import re
import sys
import unicodedata
import urllib.error
import urllib.request
from datetime import date, datetime
from pathlib import Path

SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://tkxkrbdvcctoajuigvvv.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")
PASTA_PADRAO = "dados-kalodata"
FONTE = "kalodata"

# ── Mapeamento de colunas ────────────────────────────────────────────────
# O Kalodata muda rótulo conforme o idioma e a tela. Em vez de fixar um nome,
# a gente normaliza o cabeçalho e procura por apelido. Cabeçalho não reconhecido
# não quebra o import: entra em colunas_ignoradas pra ajustar aqui depois.

ALIASES_LOJA = {
    "loja":            ["loja", "informacoes da loja", "nome da loja", "marca", "marcas e lojas",
                        "vendedor", "seller", "shop", "shop name", "store", "store name"],
    "loja_id":         ["id da loja", "shop id", "seller id", "store id"],
    "categoria":       ["categoria", "categorias", "category", "categoria principal"],
    "gmv":             ["receita", "gmv", "faturamento", "revenue", "valor de vendas",
                        "sales amount", "receita total"],
    "unidades":        ["itens vendidos", "unidades", "unidades vendidas", "vendas",
                        "units sold", "sold", "quantidade vendida"],
    "preco_medio":     ["preco medio por unidade", "preco medio", "ticket medio",
                        "average price", "avg price"],
    "produtos_ativos": ["numero de produtos", "produtos", "produtos ativos", "products",
                        "product count", "qtd produtos"],
    "criadores":       ["numero de criadores", "criadores", "criador", "influenciadores",
                        "creators", "creator count"],
    "videos":          ["numero de videos", "videos", "video count"],
    "crescimento_pct": ["taxa de crescimento da receita", "crescimento da receita",
                        "crescimento", "growth", "gmv growth", "variacao"],
    "link":            ["link", "url", "link da loja", "shop url"],
}

ALIASES_PRODUTO = {
    "produto":             ["informacoes do produto", "produto", "nome do produto",
                            "product", "product name", "titulo", "title", "item"],
    "produto_id":          ["id do produto", "id", "product id", "item id", "sku"],
    "loja":                ["loja", "informacoes da loja", "nome da loja", "marca",
                            "vendedor", "seller", "shop", "shop name", "store"],
    "categoria":           ["categoria", "categorias", "category"],
    "preco":               ["preco", "price", "preco de venda", "valor"],
    "preco_medio":         ["preco medio por unidade", "preco medio", "ticket medio",
                            "average price per unit", "average price"],
    "gmv":                 ["receita", "gmv", "faturamento", "revenue", "valor de vendas",
                            "sales amount"],
    "receita_live":        ["receita ao vivo", "receita da live", "receita live",
                            "live revenue", "livestream revenue"],
    "receita_video":       ["receita de video", "receita do video", "video revenue"],
    "receita_cartao":      ["cartao de produto", "receita do cartao de produto",
                            "product card", "product card revenue"],
    "taxa_conversao":      ["taxa de conversao do criador", "taxa de conversao",
                            "conversion rate", "creator conversion rate"],
    "unidades":            ["itens vendidos", "unidades", "unidades vendidas", "vendas",
                            "units sold", "sold", "quantidade vendida"],
    "comissao_percentual": ["taxa de comissao", "comissao", "commission",
                            "commission rate", "taxa de comissao %"],
    "comissao_reais":      ["comissao em reais", "valor da comissao", "commission amount"],
    "avaliacao":           ["avaliacao do produto", "avaliacao", "nota", "rating",
                            "review score", "estrelas"],
    "criadores":           ["numero de criadores", "criadores", "influenciadores",
                            "creators", "creator count"],
    "videos":              ["numero de videos", "videos", "video count"],
    "crescimento_pct":     ["taxa de crescimento da receita", "crescimento da receita",
                            "crescimento", "growth", "gmv growth", "variacao"],
    "link":                ["link", "url", "link do produto", "product url"],
}

# Export da aba "Vídeo e Ads" — um vídeo por linha, com receita e itens vendidos
ALIASES_VIDEO = {
    "titulo":         ["conteudo de video", "conteudo do video", "video", "titulo",
                        "legenda", "video content", "caption"],
    "arroba":         ["criador", "perfil", "conta", "creator", "username", "handle"],
    "link_video":     ["link", "url", "link do video", "video url"],
    "receita":        ["receita", "gmv", "faturamento", "revenue"],
    "views":          ["visualizacoes", "views", "view count", "reproducoes"],
    "itens_vendidos": ["itens vendidos", "unidades", "unidades vendidas", "units sold",
                        "sold"],
    "likes":          ["curtidas", "likes", "like count"],
    "comentarios":    ["comentarios", "comments", "comment count"],
    "duracao_seg":    ["duracao", "duration", "duracao do video"],
    "data_postagem":  ["data de publicacao", "data de postagem", "publicado em",
                        "publish date", "post date"],
    "produto_citado": ["produto", "nome do produto", "product", "product name"],
}

CAMPOS_NUMERO = {"gmv", "preco", "preco_medio", "comissao_reais", "receita",
                 "receita_live", "receita_video", "receita_cartao"}
CAMPOS_INTEIRO = {"unidades", "produtos_ativos", "criadores", "videos", "ranking",
                  "views", "itens_vendidos", "likes", "comentarios", "duracao_seg"}
CAMPOS_PERCENTUAL = {"crescimento_pct", "comissao_percentual", "taxa_conversao"}
CAMPOS_DECIMAL = {"avaliacao"}
CAMPOS_DATA = {"data_postagem"}
CAMPOS_BOOLEANO = {"anuncio"}


def normalizar(texto):
    """Cabeçalho vira minúsculo, sem acento e sem pontuação, pra casar com os apelidos."""
    t = unicodedata.normalize("NFD", str(texto or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn").lower()
    t = t.replace("r$", " ").replace("%", " % ")
    t = re.sub(r"[^a-z0-9%]+", " ", t)
    return re.sub(r"\s+", " ", t).strip()


# Cabeçalho do Kalodata costuma carregar o período junto: "GMV (7 dias)",
# "Tendência da receita (30/08 ~ 05/09)". A gente tira esse rabo antes de casar,
# senão "Receita ao vivo" acabaria virando "Receita".
RE_SUFIXO_PERIODO = re.compile(
    r"(?:\b(?:ultimos?|ultimas?|last)\b)?\s*\d+\s*(?:dias?|d|days?|day)?\s*$")


def sem_periodo(chave):
    anterior = None
    while anterior != chave:
        anterior = chave
        chave = RE_SUFIXO_PERIODO.sub("", chave).strip()
    return chave


def montar_mapa(cabecalhos, aliases):
    """Casa cada cabeçalho do arquivo com um campo da tabela."""
    reverso = {}
    for campo, apelidos in aliases.items():
        for apelido in apelidos:
            reverso.setdefault(normalizar(apelido), campo)

    mapa, ignoradas = {}, []
    for bruto in cabecalhos:
        chave = normalizar(bruto)
        if not chave:
            continue
        campo = reverso.get(chave) or reverso.get(sem_periodo(chave))
        if campo and campo not in mapa.values():
            mapa[bruto] = campo
        elif campo is None:
            ignoradas.append(bruto)
    return mapa, ignoradas


MULTIPLICADORES = {
    "k": 1_000, "mil": 1_000,
    "m": 1_000_000, "mi": 1_000_000, "mm": 1_000_000, "milhao": 1_000_000,
    "milhoes": 1_000_000, "b": 1_000_000_000, "bi": 1_000_000_000,
}


def para_numero(valor):
    """Entende 'R$ 1.234,56', '1,2 mil', '3.4K', '12,5%', '1.234'."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    bruto = str(valor).strip()
    if not bruto or bruto in {"-", "--", "N/A", "n/a", "—"}:
        return None
    limpo = bruto.replace("R$", "").replace("r$", "").replace("%", "").strip()
    negativo = limpo.startswith("-")
    limpo = limpo.lstrip("+-").strip()

    sufixo = ""
    achado = re.search(r"([a-zA-Zãõç]+)\s*$", limpo)
    if achado:
        candidato = normalizar(achado.group(1)).replace(" ", "")
        if candidato in MULTIPLICADORES:
            sufixo = candidato
            limpo = limpo[: achado.start()].strip()

    somente = re.sub(r"[^\d.,]", "", limpo)
    if not somente:
        return None
    # Decide quem é separador decimal: o último que aparecer manda.
    if "," in somente and "." in somente:
        if somente.rfind(",") > somente.rfind("."):
            somente = somente.replace(".", "").replace(",", ".")
        else:
            somente = somente.replace(",", "")
    elif "," in somente:
        inteiro, _, resto = somente.rpartition(",")
        somente = f"{inteiro.replace(',', '')}.{resto}" if len(resto) in (1, 2) \
            else somente.replace(",", "")
    elif somente.count(".") > 1:
        somente = somente.replace(".", "")
    elif "." in somente:
        _, _, resto = somente.rpartition(".")
        if len(resto) == 3:  # 1.234 é mil, não 1,234
            somente = somente.replace(".", "")

    try:
        numero = float(somente)
    except ValueError:
        return None
    if sufixo:
        numero *= MULTIPLICADORES[sufixo]
    return -numero if negativo else numero


def para_data(valor):
    """Aceita 18/08/2026, 2026-08-18 e 18-08-2026."""
    bruto = str(valor or "").strip()
    if not bruto:
        return None
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(bruto[:10], formato).date().isoformat()
        except ValueError:
            continue
    return None


def converter(campo, valor):
    if campo in CAMPOS_DATA:
        return para_data(valor)
    if campo in CAMPOS_BOOLEANO:
        return str(valor).strip().lower() in {"ad", "sim", "true", "1", "yes"}
    numero = para_numero(valor)
    if campo in CAMPOS_INTEIRO:
        return int(round(numero)) if numero is not None else None
    if campo in CAMPOS_NUMERO or campo in CAMPOS_PERCENTUAL or campo in CAMPOS_DECIMAL:
        return round(numero, 4) if numero is not None else None
    texto = str(valor).strip() if valor is not None else ""
    return texto or None


def ler_linhas(caminho):
    """Lê CSV (detectando o separador) ou XLSX, se openpyxl estiver disponível."""
    p = Path(caminho)
    if p.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise SystemExit(
                f"{p.name} é Excel e o openpyxl não está instalado.\n"
                "Rode: pip install openpyxl — ou exporte como CSV no Kalodata."
            )
        aba = load_workbook(p, read_only=True, data_only=True).active
        linhas = list(aba.iter_rows(values_only=True))
        if not linhas:
            return [], []
        cabecalho = [str(c) if c is not None else "" for c in linhas[0]]
        return cabecalho, [dict(zip(cabecalho, linha)) for linha in linhas[1:]]

    bruto = p.read_text(encoding="utf-8-sig", errors="replace")
    amostra = bruto[:4096]
    try:
        dialeto = csv.Sniffer().sniff(amostra, delimiters=",;\t")
        separador = dialeto.delimiter
    except csv.Error:
        separador = ";" if amostra.count(";") > amostra.count(",") else ","
    leitor = csv.DictReader(io.StringIO(bruto), delimiter=separador)
    return list(leitor.fieldnames or []), list(leitor)


def detectar_tipo(cabecalhos):
    chaves = {sem_periodo(normalizar(c)) for c in cabecalhos}
    if "conteudo de video" in chaves or ("visualizacoes" in chaves and
                                         "data de publicacao" in chaves):
        return "videos"
    tem_produto = any(k in chaves for k in
                      [normalizar(a) for a in ALIASES_PRODUTO["produto"]])
    if tem_produto:
        return "produtos"
    tem_loja = any(k in chaves for k in [normalizar(a) for a in ALIASES_LOJA["loja"]])
    if tem_loja:
        return "lojas"
    return None


def http(metodo, caminho, corpo=None, headers=None):
    if not SUPABASE_KEY:
        raise SystemExit("Falta a variável de ambiente SUPABASE_KEY.")
    cabecalhos = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
    }
    cabecalhos.update(headers or {})
    dados = json.dumps(corpo).encode() if corpo is not None else None
    req = urllib.request.Request(f"{SUPABASE_URL}/rest/v1/{caminho}",
                                 data=dados, headers=cabecalhos, method=metodo)
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            texto = r.read().decode()
            return json.loads(texto) if texto.strip() else []
    except urllib.error.HTTPError as e:
        raise SystemExit(f"Supabase respondeu {e.code}: {e.read().decode()[:500]}")


def gravar(tabela, registros, conflito):
    if not registros:
        return 0
    salvos = 0
    for i in range(0, len(registros), 200):
        lote = registros[i:i + 200]
        http("POST", f"{tabela}?on_conflict={conflito}", lote,
             {"Prefer": "resolution=merge-duplicates,return=minimal"})
        salvos += len(lote)
    return salvos


def gravar_simples(tabela, registros):
    """Insert puro, pra linha que não tem chave de upsert (vídeo sem link)."""
    if not registros:
        return 0
    for i in range(0, len(registros), 200):
        http("POST", tabela, registros[i:i + 200], {"Prefer": "return=minimal"})
    return len(registros)


def importar(caminho, periodo, data_ref, dry_run, produto_alvo=None):
    cabecalhos, linhas = ler_linhas(caminho)
    if not linhas:
        print(f"  {Path(caminho).name}: arquivo vazio, pulei.")
        return

    tipo = detectar_tipo(cabecalhos)
    if tipo is None:
        print(f"  {Path(caminho).name}: não achei coluna de produto nem de loja.")
        print(f"     Cabeçalhos lidos: {cabecalhos}")
        return

    config = {
        "produtos": (ALIASES_PRODUTO, "tiktok_ranking_produtos",
                     "produto,loja,periodo,data_ref", "produto"),
        "lojas":    (ALIASES_LOJA, "tiktok_lojas", "loja,periodo,data_ref", "loja"),
        "videos":   (ALIASES_VIDEO, "tiktok_videos", "link_video", "titulo"),
    }
    aliases, tabela, conflito, obrigatorio = config[tipo]

    mapa, ignoradas = montar_mapa(cabecalhos, aliases)
    print(f"  {Path(caminho).name} → {tabela} ({len(linhas)} linha(s))")
    for bruto, campo in mapa.items():
        print(f"     {bruto!r} → {campo}")
    if ignoradas:
        print(f"     não mapeadas: {ignoradas}")

    registros = []
    for i, linha in enumerate(linhas, start=1):
        reg = {"periodo": periodo, "data_ref": data_ref, "fonte": FONTE}
        if tipo != "videos":
            reg["ranking"] = i
        for bruto, campo in mapa.items():
            reg[campo] = converter(campo, linha.get(bruto))
        if not reg.get(obrigatorio):
            continue
        if tipo == "videos":
            bruto_titulo = next((b for b, c in mapa.items() if c == "titulo"), None)
            texto = str(linha.get(bruto_titulo) or "")
            reg["anuncio"] = converter("anuncio", "ad") if re.search(
                r"\bads?\b", texto, re.I) else False
            # tiktok_videos guarda a legenda no gancho; titulo não é coluna de lá
            reg["gancho"] = reg.pop("titulo", None)
            reg["data_deteccao"] = data_ref
            if produto_alvo:
                reg.setdefault("produto_citado", produto_alvo)
            if not reg.get("link_video"):
                # sem link não dá pra fazer upsert; o índice único só cobre quem tem link
                reg.pop("link_video", None)
        registros.append(reg)

    if tipo == "videos" and any(not r.get("link_video") for r in registros):
        print("     aviso: linhas sem link de vídeo entram como novas a cada import.")

    print(f"     {len(registros)} linha(s) válida(s) de {len(linhas)}")
    if registros:
        print(f"     exemplo: {json.dumps(registros[0], ensure_ascii=False)[:300]}")

    if dry_run:
        print("     (dry-run: nada foi gravado)")
        return

    com_chave = [r for r in registros if tipo != "videos" or r.get("link_video")]
    sem_chave = [r for r in registros if tipo == "videos" and not r.get("link_video")]
    salvos = gravar(tabela, com_chave, conflito) + gravar_simples(tabela, sem_chave)
    http("POST", "tiktok_importacoes", [{
        "arquivo": Path(caminho).name, "tipo": tipo, "periodo": periodo,
        "data_ref": data_ref, "linhas_lidas": len(linhas), "linhas_salvas": salvos,
        "colunas_ignoradas": ignoradas,
    }], {"Prefer": "return=minimal"})
    print(f"     {salvos} linha(s) gravada(s) no Supabase.")


def main():
    args = sys.argv[1:]
    dry_run = "--dry-run" in args
    args = [a for a in args if a != "--dry-run"]

    periodo, data_ref, produto_alvo = "7d", date.today().isoformat(), None
    restantes = []
    i = 0
    while i < len(args):
        if args[i] == "--periodo" and i + 1 < len(args):
            periodo = args[i + 1]; i += 2
        elif args[i] == "--produto" and i + 1 < len(args):
            produto_alvo = args[i + 1]; i += 2
        elif args[i] == "--data" and i + 1 < len(args):
            data_ref = datetime.strptime(args[i + 1], "%Y-%m-%d").date().isoformat()
            i += 2
        else:
            restantes.append(args[i]); i += 1

    if periodo not in {"1d", "7d", "30d", "90d"}:
        raise SystemExit("--periodo aceita 1d, 7d, 30d ou 90d.")

    arquivos = restantes or sorted(
        str(p) for p in Path(PASTA_PADRAO).glob("*")
        if p.suffix.lower() in {".csv", ".xlsx", ".xlsm"}
    )
    if not arquivos:
        raise SystemExit(
            f"Nenhum arquivo pra importar. Coloque o export do Kalodata em {PASTA_PADRAO}/ "
            "ou passe o caminho: python tiktok-sync.py export.csv"
        )

    print(f"Importando período {periodo}, data de referência {data_ref}")
    for arquivo in arquivos:
        importar(arquivo, periodo, data_ref, dry_run, produto_alvo)
    print("Pronto.")


if __name__ == "__main__":
    main()
